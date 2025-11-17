"""
XLSX to text converter.
"""
from openpyxl import load_workbook
from .base import BaseConverter


class XLSXConverter(BaseConverter):
    """Converter for XLSX spreadsheets."""

    def convert(self, input_path: str, output_path: str) -> bool:
        """
        Convert XLSX to text.

        Args:
            input_path: Path to XLSX file
            output_path: Path to output text file

        Returns:
            True if successful, False otherwise
        """
        try:
            text = self._extract_text(input_path)
            if text:
                return self._write_output(text, output_path)
            return False
        except Exception as e:
            self.logger.error(f"Error converting XLSX {input_path}: {str(e)}")
            return False

    def _extract_text(self, input_path: str) -> str:
        """
        Extract text from XLSX.

        Args:
            input_path: Path to XLSX file

        Returns:
            Extracted text
        """
        workbook = load_workbook(input_path, data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if self.preserve_structure:
                text_parts.append(f"\n{'='*60}\n")
                text_parts.append(f"Sheet: {sheet_name}\n")
                text_parts.append(f"{'='*60}\n\n")

            # Get all rows
            for row in sheet.iter_rows(values_only=True):
                # Filter out empty cells
                row_values = [str(cell) if cell is not None else '' for cell in row]
                # Skip completely empty rows
                if any(val.strip() for val in row_values):
                    if self.preserve_structure:
                        text_parts.append(" | ".join(row_values) + "\n")
                    else:
                        text_parts.append(" ".join(row_values) + "\n")

            text_parts.append("\n")

        return ''.join(text_parts).strip()

    @classmethod
    def get_supported_extensions(cls) -> list[str]:
        """Get supported file extensions."""
        return ['.xlsx']
